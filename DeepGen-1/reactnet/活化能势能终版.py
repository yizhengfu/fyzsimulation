import os
import re
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, simpledialog, colorchooser, scrolledtext
from openpyxl import load_workbook
import sys
import time
import numpy as np
from scipy.optimize import curve_fit
from scipy.stats import linregress
import matplotlib.pyplot as plt
from matplotlib.font_manager import FontProperties
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from PIL import Image, ImageTk
import openpyxl.drawing.image
import matplotlib as mpl
import tkinter.font as tkFont
import json
import os.path

# 设置Matplotlib使用支持中文的字体
plt.rcParams['font.sans-serif'] = ['SimHei', 'Arial Unicode MS', 'Microsoft YaHei', 'DejaVu Sans']  # 使用支持中文的字体
plt.rcParams['axes.unicode_minus'] = False  # 正确显示负号

# 定义全局变量用于指数拟合
a = 0.0
b = 0.0
c = 0.0
include_c_point = True  # 默认包含X=c的数据点

# 理想气体常数 (J/mol·K)
R = 8.314

# 设置文件路径
SETTINGS_FILE = "plot_settings.json"

# 默认设置
DEFAULT_ACTIVATION_SETTINGS = {
    'scatter_size': 50,
    'scatter_color': 'blue',
    'line_color': 'red',
    'line_width': 2,
    'axis_title_fontsize': 12,
    'axis_tick_fontsize': 10,
    'show_legend': True,
    'legend_frame': False,  # 新增：图例无边框
    'show_statistics': True,
    'x_label': '1000/T (K⁻¹)',
    'y_label': 'ln(k)',
    'legend_label_data': '原始数据点',
    'legend_label_fit': '拟合直线',
    'statistics_fontsize': 10,
    'statistics_position_x': 0.05,
    'statistics_position_y': 0.95,
    'show_grid': True,
    # 独立字体设置
    'font_family': 'Times New Roman',
    'title_font_family': 'Times New Roman',
    'legend_font_family': 'Times New Roman',
    'axis_font_family': 'Times New Roman',
    'tick_font_family': 'Times New Roman',
    'font_size': 10,
    'title_font_size': 14,
    'legend_font_size': 10,
    'axis_font_size': 12,
    'tick_font_size': 10
}

DEFAULT_MULTI_TEMP_SETTINGS = {
    'line_width': 2,
    'axis_title_fontsize': 12,
    'axis_tick_fontsize': 10,
    'show_legend': True,
    'legend_frame': False,  # 新增：图例无边框
    'x_label': 'Time (ps)',
    'y_label': 'Potential Energy',
    'show_grid': True,
    'show_title': False,  # 新增：是否显示标题
    'title_text': '不同温度下能量随时间变化',  # 新增：标题文本
    # 独立字体设置
    'font_family': 'Times New Roman',
    'title_font_family': 'Times New Roman',
    'legend_font_family': 'Times New Roman',
    'axis_font_family': 'Times New Roman',
    'tick_font_family': 'Times New Roman',
    'font_size': 10,
    'title_font_size': 14,
    'legend_font_size': 10,
    'axis_font_size': 12,
    'tick_font_size': 10
}

# 全局变量用于存储绘图设置
activation_plot_settings = DEFAULT_ACTIVATION_SETTINGS.copy()
multi_temp_plot_settings = DEFAULT_MULTI_TEMP_SETTINGS.copy()


# 加载保存的设置
def load_settings():
    global activation_plot_settings, multi_temp_plot_settings
    if os.path.exists(SETTINGS_FILE):
        try:
            with open(SETTINGS_FILE, 'r') as f:
                settings = json.load(f)
                activation_plot_settings = settings.get('activation', DEFAULT_ACTIVATION_SETTINGS.copy())
                multi_temp_plot_settings = settings.get('multi_temp', DEFAULT_MULTI_TEMP_SETTINGS.copy())
        except:
            # 如果加载失败，使用默认设置
            activation_plot_settings = DEFAULT_ACTIVATION_SETTINGS.copy()
            multi_temp_plot_settings = DEFAULT_MULTI_TEMP_SETTINGS.copy()
    else:
        activation_plot_settings = DEFAULT_ACTIVATION_SETTINGS.copy()
        multi_temp_plot_settings = DEFAULT_MULTI_TEMP_SETTINGS.copy()


# 保存设置
def save_settings():
    settings = {
        'activation': activation_plot_settings,
        'multi_temp': multi_temp_plot_settings
    }
    with open(SETTINGS_FILE, 'w') as f:
        json.dump(settings, f, indent=4)


# 初始化时加载设置
load_settings()


def exponential_func(x, k):
    """
    指数函数 Y = a + b * exp[-k*(X - c)]
    其中a, b, c是已知参数，k是待拟合参数
    """
    global a, b, c
    return a + b * np.exp(-k * (x - c))


def process_log_file(file_path, preview_text=None):
    """
    处理日志文件，提取两个关键内容之间的数据并保存为Excel文件
    """
    try:
        if preview_text:
            preview_text.insert(tk.END, f"处理文件: {os.path.basename(file_path)}\n")
            preview_text.see(tk.END)
            preview_text.update_idletasks()

        # 读取文件内容
        with open(file_path, 'r', encoding='utf-8', errors='replace') as file:
            content = file.read()

        # 尝试多种可能的模式匹配
        patterns = [
            r'Per MPI rank memory allocation(.*?)Loop time of',  # 原始模式
            r'(Step\s+Temp\s+PotEng\s+KinEng\s+TotEng\s+Press.*?Loop time of)',  # 带列标题的模式
            r'(Step[\s\S]*?Loop time of)',  # 更通用的模式
            r'(LAMMPS\s+\([\s\S]*?)Loop time of',  # 以LAMMPS开头的模式
            r'(Step.*?\n)(.*?)(?=\n\s*\n|\Z)',  # 匹配数据块直到空行或文件结束
            r'(Step.*?)\n(.*?)(?=Loop time of|\Z)',  # 匹配数据块直到"Loop time of"或文件结束
        ]

        extracted_data = None

        for pattern in patterns:
            match = re.search(pattern, content, re.DOTALL)
            if match:
                extracted_data = match.group(1).strip()
                break

        if not extracted_data:
            # 尝试查找包含列标题的数据块
            header_pattern = r'(Step\s+[^\n]+\n)'
            header_match = re.search(header_pattern, content)
            if header_match:
                header = header_match.group(1)
                # 查找从标题开始到文件结束或空行的数据
                data_pattern = re.escape(header) + r'(.*?)(?=\n\s*\n|\Z)'
                data_match = re.search(data_pattern, content, re.DOTALL)
                if data_match:
                    extracted_data = header + data_match.group(1)

        if not extracted_data:
            # 尝试查找包含"Step"的数据块
            step_pattern = r'(Step.*?)\n(.*?)(?=\n\s*\n|\Z)'
            step_match = re.search(step_pattern, content, re.DOTALL)
            if step_match:
                extracted_data = step_match.group(0).strip()

        if not extracted_data:
            # 最后尝试：查找所有看起来像数据行的内容
            data_lines = []
            for line in content.split('\n'):
                # 匹配包含数字的行（假设至少包含4个数值）
                if re.match(r'^\s*\d+\s+-?\d+\.\d+\s+-?\d+\.\d+\s+-?\d+\.\d+', line):
                    data_lines.append(line)

            if data_lines:
                extracted_data = "\n".join(data_lines)
            else:
                # 如果仍然找不到，尝试提取所有数值行
                data_lines = []
                for line in content.split('\n'):
                    # 匹配包含至少4个数值的行
                    if len(re.findall(r'-?\d+\.\d+', line)) >= 4:
                        data_lines.append(line)

                if data_lines:
                    extracted_data = "\n".join(data_lines)
                else:
                    error_msg = f"未找到可识别的数据格式: {os.path.basename(file_path)}"
                    if preview_text:
                        preview_text.insert(tk.END, f"错误: {error_msg}\n", "error")
                        preview_text.see(tk.END)
                        preview_text.update_idletasks()
                    raise ValueError(error_msg)

        # 处理提取的数据
        processed_data = []
        lines = extracted_data.split('\n')

        # 处理标题行（如果存在）
        header_line = lines[0] if any(char.isalpha() for char in lines[0]) else None
        if header_line:
            # 处理标题行
            headers = re.split(r'\s+', header_line.strip())
            processed_data.append(headers)
            start_index = 1
        else:
            start_index = 0

        # 处理数据行
        for line in lines[start_index:]:
            if not line.strip():
                continue

            parts = re.split(r'\s+', line.strip())
            # 如果第一个元素是空字符串，则移除
            if parts and parts[0] == '':
                parts = parts[1:]
            processed_data.append(parts)

        # 创建输出文件
        output_file = os.path.splitext(file_path)[0] + '.xlsx'

        # 将处理后的数据转换为DataFrame
        if not processed_data:
            error_msg = "提取的数据为空，请检查日志文件格式"
            if preview_text:
                preview_text.insert(tk.END, f"错误: {error_msg}\n", "error")
                preview_text.see(tk.END)
                preview_text.update_idletasks()
            raise ValueError(error_msg)

        # 确保所有行长度一致
        max_cols = max(len(row) for row in processed_data)
        for row in processed_data:
            if len(row) < max_cols:
                row.extend([''] * (max_cols - len(row)))

        df = pd.DataFrame(processed_data)

        # 保存为Excel文件
        df.to_excel(output_file, index=False, header=False)

        if preview_text:
            preview_text.insert(tk.END, f"成功创建: {os.path.basename(output_file)}\n", "success")
            preview_text.see(tk.END)
            preview_text.update_idletasks()

            # 预览前5行数据
            preview_text.insert(tk.END, "数据预览:\n")
            for i in range(min(5, len(processed_data))):
                preview_text.insert(tk.END, f"{processed_data[i]}\n")
            preview_text.insert(tk.END, "\n")
            preview_text.see(tk.END)
            preview_text.update_idletasks()

        return output_file

    except Exception as e:
        error_msg = f"处理文件时出错: {str(e)}"
        if preview_text:
            preview_text.insert(tk.END, f"错误: {error_msg}\n", "error")
            preview_text.see(tk.END)
            preview_text.update_idletasks()
        raise RuntimeError(error_msg) from e


def find_log_files(folder_path, preview_text=None):
    """查找所有匹配的日志文件"""
    log_files = []
    if preview_text:
        preview_text.insert(tk.END, f"搜索文件夹: {folder_path}\n")
        preview_text.see(tk.END)
        preview_text.update_idletasks()

    for root, _, files in os.walk(folder_path):
        if preview_text:
            preview_text.insert(tk.END, f"扫描子文件夹: {os.path.basename(root)}\n")
            preview_text.see(tk.END)
            preview_text.update_idletasks()

        for file in files:
            # 匹配多种可能的日志文件名
            if (re.match(r'log\.\d+decom\.txt$', file) or
                    re.match(r'log\.lammps$', file) or
                    re.match(r'log\.\d+\.txt$', file) or
                    re.match(r'log\.txt$', file) or
                    re.match(r'log.*\.txt$', file) or
                    file.lower().startswith('log') and file.lower().endswith(('.txt', '.log'))):
                log_files.append(os.path.join(root, file))
                if preview_text:
                    preview_text.insert(tk.END, f"找到日志文件: {file}\n")
                    preview_text.see(tk.END)
                    preview_text.update_idletasks()

    if preview_text:
        preview_text.insert(tk.END, f"共找到 {len(log_files)} 个日志文件\n\n")
        preview_text.see(tk.END)
        preview_text.update_idletasks()

    return log_files


def choose_font_dialog(parent, current_family, current_size, title="选择字体"):
    """创建字体选择对话框"""
    font_window = tk.Toplevel(parent)
    font_window.title(title)
    font_window.geometry("400x400")
    font_window.transient(parent)
    font_window.grab_set()

    # 获取可用字体
    available_fonts = list(tkFont.families())
    available_fonts.sort()

    # 字体列表框架
    font_frame = ttk.Frame(font_window)
    font_frame.pack(fill="both", expand=True, padx=10, pady=10)

    # 字体列表
    font_list = tk.Listbox(font_frame, width=30)
    font_scroll = ttk.Scrollbar(font_frame, orient="vertical", command=font_list.yview)
    font_list.config(yscrollcommand=font_scroll.set)

    # 填充字体列表
    for font in available_fonts:
        font_list.insert(tk.END, font)

    # 设置当前选中项
    if current_family in available_fonts:
        idx = available_fonts.index(current_family)
        font_list.selection_set(idx)
        font_list.see(idx)

    # 字体大小
    ttk.Label(font_window, text="字体大小:").pack(pady=(10, 0))
    size_var = tk.StringVar(value=str(current_size))
    size_entry = ttk.Entry(font_window, textvariable=size_var, width=10)
    size_entry.pack()

    # 预览标签
    preview_label = ttk.Label(font_window, text="字体预览 ABCabc123",
                              font=(current_family, current_size))
    preview_text = tk.Label(font_window, text="字体预览 ABCabc123",
                            font=(current_family, current_size))
    preview_label.pack(pady=10)

    # 更新预览
    def update_preview(event=None):
        selected = font_list.curselection()
        if selected:
            family = font_list.get(selected[0])
            try:
                size = int(size_var.get())
                preview_label.config(font=(family, size))
                preview_text.config(font=(family, size))
            except:
                pass

    font_list.bind("<<ListboxSelect>>", update_preview)
    size_entry.bind("<KeyRelease>", update_preview)

    # 结果变量
    result = {"family": current_family, "size": current_size}

    # 确认按钮
    def confirm_font():
        selected = font_list.curselection()
        if selected:
            family = font_list.get(selected[0])
            try:
                size = int(size_var.get())
                result["family"] = family
                result["size"] = size
                font_window.destroy()
                return result
            except ValueError:
                messagebox.showerror("错误", "请输入有效的字体大小")
        return None

    # 取消按钮
    def cancel_font():
        font_window.destroy()
        return None

    btn_frame = tk.Frame(font_window)
    btn_frame.pack(pady=10)

    ttk.Button(btn_frame, text="确认", command=confirm_font).pack(side=tk.LEFT, padx=10)
    ttk.Button(btn_frame, text="取消", command=cancel_font).pack(side=tk.LEFT, padx=10)

    font_list.pack(side="left", fill="both", expand=True)
    font_scroll.pack(side="right", fill="y")

    # 等待窗口关闭
    font_window.wait_window()

    return result


def get_activation_plot_settings():
    """获取用户自定义的活化能图绘图设置"""
    global activation_plot_settings

    # 创建设置窗口
    settings_window = tk.Toplevel()
    settings_window.title("活化能图设置")
    settings_window.geometry("800x900")  # 增加窗口大小以适应更多设置

    # 创建主框架
    main_frame = ttk.Frame(settings_window)
    main_frame.pack(fill="both", expand=True, padx=10, pady=10)

    # 创建画布和滚动条
    canvas = tk.Canvas(main_frame)
    scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
    scrollable_frame = ttk.Frame(canvas)

    scrollable_frame.bind(
        "<Configure>",
        lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
    )

    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)

    canvas.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")

    # 创建标签框架
    scatter_frame = ttk.LabelFrame(scrollable_frame, text="散点设置")
    scatter_frame.pack(fill="x", padx=10, pady=5)

    line_frame = ttk.LabelFrame(scrollable_frame, text="直线设置")
    line_frame.pack(fill="x", padx=10, pady=5)

    axis_frame = ttk.LabelFrame(scrollable_frame, text="坐标轴设置")
    axis_frame.pack(fill="x", padx=10, pady=5)

    legend_frame = ttk.LabelFrame(scrollable_frame, text="图例设置")
    legend_frame.pack(fill="x", padx=10, pady=5)

    stats_frame = ttk.LabelFrame(scrollable_frame, text="统计信息设置")
    stats_frame.pack(fill="x", padx=10, pady=5)

    options_frame = ttk.LabelFrame(scrollable_frame, text="其他选项")
    options_frame.pack(fill="x", padx=10, pady=5)

    # 新增：字体设置框架
    font_frame = ttk.LabelFrame(scrollable_frame, text="字体设置")
    font_frame.pack(fill="x", padx=10, pady=5)

    # 散点设置
    ttk.Label(scatter_frame, text="散点大小:").grid(row=0, column=0, padx=5, pady=5, sticky="e")
    scatter_size_var = tk.StringVar(value=str(activation_plot_settings['scatter_size']))
    scatter_size_entry = ttk.Entry(scatter_frame, textvariable=scatter_size_var, width=10)
    scatter_size_entry.grid(row=0, column=1, padx=5, pady=5, sticky="w")

    ttk.Label(scatter_frame, text="散点颜色:").grid(row=1, column=0, padx=5, pady=5, sticky="e")
    scatter_color_var = tk.StringVar(value=activation_plot_settings['scatter_color'])
    scatter_color_entry = ttk.Entry(scatter_frame, textvariable=scatter_color_var, width=10)
    scatter_color_entry.grid(row=1, column=1, padx=5, pady=5, sticky="w")

    def choose_scatter_color():
        color = colorchooser.askcolor(title="选择散点颜色")
        if color[1]:
            scatter_color_var.set(color[1])
            scatter_color_entry.config(background=color[1])

    ttk.Button(scatter_frame, text="选择颜色", command=choose_scatter_color).grid(row=1, column=2, padx=5, pady=5)

    # 直线设置
    ttk.Label(line_frame, text="直线颜色:").grid(row=0, column=0, padx=5, pady=5, sticky="e")
    line_color_var = tk.StringVar(value=activation_plot_settings['line_color'])
    line_color_entry = ttk.Entry(line_frame, textvariable=line_color_var, width=10)
    line_color_entry.grid(row=0, column=1, padx=5, pady=5, sticky="w")

    def choose_line_color():
        color = colorchooser.askcolor(title="选择直线颜色")
        if color[1]:
            line_color_var.set(color[1])
            line_color_entry.config(background=color[1])

    ttk.Button(line_frame, text="选择颜色", command=choose_line_color).grid(row=0, column=2, padx=5, pady=5)

    ttk.Label(line_frame, text="直线宽度:").grid(row=1, column=0, padx=5, pady=5, sticky="e")
    line_width_var = tk.StringVar(value=str(activation_plot_settings['line_width']))
    line_width_entry = ttk.Entry(line_frame, textvariable=line_width_var, width=10)
    line_width_entry.grid(row=1, column=1, padx=5, pady=5, sticky="w")

    # 坐标轴设置
    ttk.Label(axis_frame, text="X轴标题:").grid(row=0, column=0, padx=5, pady=5, sticky="e")
    x_label_var = tk.StringVar(value=activation_plot_settings['x_label'])
    x_label_entry = ttk.Entry(axis_frame, textvariable=x_label_var, width=30)
    x_label_entry.grid(row=0, column=1, padx=5, pady=5, sticky="w")

    ttk.Label(axis_frame, text="Y轴标题:").grid(row=1, column=0, padx=5, pady=5, sticky="e")
    y_label_var = tk.StringVar(value=activation_plot_settings['y_label'])
    y_label_entry = ttk.Entry(axis_frame, textvariable=y_label_var, width=30)
    y_label_entry.grid(row=1, column=1, padx=5, pady=5, sticky="w")

    ttk.Label(axis_frame, text="坐标轴标题字号:").grid(row=2, column=0, padx=5, pady=5, sticky="e")
    axis_title_fontsize_var = tk.StringVar(value=str(activation_plot_settings['axis_title_fontsize']))
    axis_title_fontsize_entry = ttk.Entry(axis_frame, textvariable=axis_title_fontsize_var, width=10)
    axis_title_fontsize_entry.grid(row=2, column=1, padx=5, pady=5, sticky="w")

    ttk.Label(axis_frame, text="坐标轴刻度字号:").grid(row=3, column=0, padx=5, pady=5, sticky="e")
    axis_tick_fontsize_var = tk.StringVar(value=str(activation_plot_settings['axis_tick_fontsize']))
    axis_tick_fontsize_entry = ttk.Entry(axis_frame, textvariable=axis_tick_fontsize_var, width=10)
    axis_tick_fontsize_entry.grid(row=3, column=1, padx=5, pady=5, sticky="w")

    # 图例设置
    ttk.Label(legend_frame, text="数据点图例文本:").grid(row=0, column=0, padx=5, pady=5, sticky="e")
    legend_data_var = tk.StringVar(value=activation_plot_settings['legend_label_data'])
    legend_data_entry = ttk.Entry(legend_frame, textvariable=legend_data_var, width=30)
    legend_data_entry.grid(row=0, column=1, padx=5, pady=5, sticky="w")

    ttk.Label(legend_frame, text="拟合直线图例文本:").grid(row=1, column=0, padx=5, pady=5, sticky="e")
    legend_fit_var = tk.StringVar(value=activation_plot_settings['legend_label_fit'])
    legend_fit_entry = ttk.Entry(legend_frame, textvariable=legend_fit_var, width=30)
    legend_fit_entry.grid(row=1, column=1, padx=5, pady=5, sticky="w")

    show_legend_var = tk.BooleanVar(value=activation_plot_settings['show_legend'])
    show_legend_cb = ttk.Checkbutton(legend_frame, text="显示图例", variable=show_legend_var)
    show_legend_cb.grid(row=2, column=0, padx=5, pady=5, sticky="w")

    # 新增：图例边框设置
    legend_frame_var = tk.BooleanVar(value=activation_plot_settings['legend_frame'])
    legend_frame_cb = ttk.Checkbutton(legend_frame, text="图例边框", variable=legend_frame_var)
    legend_frame_cb.grid(row=2, column=1, padx=5, pady=5, sticky="w")

    # 统计信息设置
    ttk.Label(stats_frame, text="统计信息位置X(0-1):").grid(row=0, column=0, padx=5, pady=5, sticky="e")
    stats_x_var = tk.StringVar(value=str(activation_plot_settings['statistics_position_x']))
    stats_x_entry = ttk.Entry(stats_frame, textvariable=stats_x_var, width=10)
    stats_x_entry.grid(row=0, column=1, padx=5, pady=5, sticky="w")

    ttk.Label(stats_frame, text="统计信息位置Y(0-1):").grid(row=1, column=0, padx=5, pady=5, sticky="e")
    stats_y_var = tk.StringVar(value=str(activation_plot_settings['statistics_position_y']))
    stats_y_entry = ttk.Entry(stats_frame, textvariable=stats_y_var, width=10)
    stats_y_entry.grid(row=1, column=1, padx=5, pady=5, sticky="w")

    ttk.Label(stats_frame, text="统计信息字号:").grid(row=2, column=0, padx=5, pady=5, sticky="e")
    stats_fontsize_var = tk.StringVar(value=str(activation_plot_settings['statistics_fontsize']))
    stats_fontsize_entry = ttk.Entry(stats_frame, textvariable=stats_fontsize_var, width=10)
    stats_fontsize_entry.grid(row=2, column=1, padx=5, pady=5, sticky="w")

    show_stats_var = tk.BooleanVar(value=activation_plot_settings['show_statistics'])
    show_stats_cb = ttk.Checkbutton(stats_frame, text="显示统计信息", variable=show_stats_var)
    show_stats_cb.grid(row=3, column=0, padx=5, pady=5, sticky="w")

    # 其他选项
    show_grid_var = tk.BooleanVar(value=activation_plot_settings['show_grid'])
    show_grid_cb = ttk.Checkbutton(options_frame, text="显示网格", variable=show_grid_var)
    show_grid_cb.grid(row=0, column=0, padx=5, pady=5, sticky="w")

    # 字体设置
    font_rows = [
        ("全局字体", "font_family", "font_size"),
        ("标题字体", "title_font_family", "title_font_size"),
        ("图例字体", "legend_font_family", "legend_font_size"),
        ("坐标轴标题字体", "axis_font_family", "axis_font_size"),
        ("刻度字体", "tick_font_family", "tick_font_size")
    ]

    # 存储字体标签和StringVar的字典
    font_labels = {}

    for i, (label, family_key, size_key) in enumerate(font_rows):
        ttk.Label(font_frame, text=f"{label}:").grid(row=i, column=0, padx=5, pady=5, sticky="e")

        # 显示当前字体设置
        current_family = activation_plot_settings[family_key]
        current_size = activation_plot_settings[size_key]
        current_font_label = f"{current_family}, {current_size}"
        font_var = tk.StringVar(value=current_font_label)
        font_label = ttk.Label(font_frame, textvariable=font_var)
        font_label.grid(row=i, column=1, padx=5, pady=5, sticky="w")

        # 存储标签和StringVar
        font_labels[(family_key, size_key)] = (font_var, font_label)

        # 字体选择按钮
        def create_choose_font(fk=family_key, sk=size_key, lbl=label):
            def choose_font():
                font_result = choose_font_dialog(
                    settings_window,
                    activation_plot_settings[fk],
                    activation_plot_settings[sk],
                    title=f"选择{lbl}字体"
                )
                if font_result:
                    activation_plot_settings[fk] = font_result["family"]
                    activation_plot_settings[sk] = font_result["size"]
                    # 更新显示
                    font_var, _ = font_labels[(fk, sk)]
                    font_var.set(f"{font_result['family']}, {font_result['size']}")

            return choose_font

        ttk.Button(font_frame, text="选择字体...",
                   command=create_choose_font()).grid(row=i, column=2, padx=5, pady=5)

    # 设置初始颜色
    scatter_color_entry.config(background=activation_plot_settings['scatter_color'])
    line_color_entry.config(background=activation_plot_settings['line_color'])

    # 确认按钮
    def confirm_settings():
        try:
            activation_plot_settings['scatter_size'] = int(scatter_size_var.get())
            activation_plot_settings['scatter_color'] = scatter_color_var.get()
            activation_plot_settings['line_color'] = line_color_var.get()
            activation_plot_settings['line_width'] = float(line_width_var.get())
            activation_plot_settings['axis_title_fontsize'] = int(axis_title_fontsize_var.get())
            activation_plot_settings['axis_tick_fontsize'] = int(axis_tick_fontsize_var.get())
            activation_plot_settings['show_legend'] = show_legend_var.get()
            activation_plot_settings['legend_frame'] = legend_frame_var.get()  # 新增
            activation_plot_settings['show_statistics'] = show_stats_var.get()
            activation_plot_settings['x_label'] = x_label_var.get()
            activation_plot_settings['y_label'] = y_label_var.get()
            activation_plot_settings['legend_label_data'] = legend_data_var.get()
            activation_plot_settings['legend_label_fit'] = legend_fit_var.get()
            activation_plot_settings['statistics_fontsize'] = int(stats_fontsize_var.get())
            activation_plot_settings['statistics_position_x'] = float(stats_x_var.get())
            activation_plot_settings['statistics_position_y'] = float(stats_y_var.get())
            activation_plot_settings['show_grid'] = show_grid_var.get()

            # 保存设置
            save_settings()
            settings_window.destroy()
        except ValueError as e:
            messagebox.showerror("输入错误", f"请输入有效的数值: {str(e)}")

    # 将确认按钮放在主框架底部，确保始终可见
    btn_frame = tk.Frame(main_frame)
    btn_frame.pack(side=tk.BOTTOM, fill=tk.X, padx=10, pady=10)

    ttk.Button(btn_frame, text="确认", command=confirm_settings).pack(pady=10)

    # 等待窗口关闭
    settings_window.wait_window()

    return activation_plot_settings


def get_multi_temp_plot_settings():
    """获取用户自定义的多温度能量图绘图设置"""
    global multi_temp_plot_settings

    # 创建设置窗口
    settings_window = tk.Toplevel()
    settings_window.title("多温度能量图设置")
    settings_window.geometry("800x900")  # 增加窗口大小以适应更多设置

    # 创建主框架
    main_frame = ttk.Frame(settings_window)
    main_frame.pack(fill="both", expand=True, padx=10, pady=10)

    # 创建画布和滚动条
    canvas = tk.Canvas(main_frame)
    scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
    scrollable_frame = ttk.Frame(canvas)

    scrollable_frame.bind(
        "<Configure>",
        lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
    )

    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)

    canvas.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")

    # 创建标签框架
    line_frame = ttk.LabelFrame(scrollable_frame, text="曲线设置")
    line_frame.pack(fill="x", padx=10, pady=5)

    axis_frame = ttk.LabelFrame(scrollable_frame, text="坐标轴设置")
    axis_frame.pack(fill="x", padx=10, pady=5)

    legend_frame = ttk.LabelFrame(scrollable_frame, text="图例设置")
    legend_frame.pack(fill="x", padx=10, pady=5)

    title_frame = ttk.LabelFrame(scrollable_frame, text="标题设置")
    title_frame.pack(fill="x", padx=10, pady=5)

    options_frame = ttk.LabelFrame(scrollable_frame, text="其他选项")
    options_frame.pack(fill="x", padx=10, pady=5)

    # 新增：字体设置框架
    font_frame = ttk.LabelFrame(scrollable_frame, text="字体设置")
    font_frame.pack(fill="x", padx=10, pady=5)

    # 曲线设置
    ttk.Label(line_frame, text="曲线宽度:").grid(row=0, column=0, padx=5, pady=5, sticky="e")
    line_width_var = tk.StringVar(value=str(multi_temp_plot_settings['line_width']))
    line_width_entry = ttk.Entry(line_frame, textvariable=line_width_var, width=10)
    line_width_entry.grid(row=0, column=1, padx=5, pady=5, sticky="w")

    # 坐标轴设置
    ttk.Label(axis_frame, text="X轴标题:").grid(row=0, column=0, padx=5, pady=5, sticky="e")
    x_label_var = tk.StringVar(value=multi_temp_plot_settings['x_label'])
    x_label_entry = ttk.Entry(axis_frame, textvariable=x_label_var, width=30)
    x_label_entry.grid(row=0, column=1, padx=5, pady=5, sticky="w")

    ttk.Label(axis_frame, text="Y轴标题:").grid(row=1, column=0, padx=5, pady=5, sticky="e")
    y_label_var = tk.StringVar(value=multi_temp_plot_settings['y_label'])
    y_label_entry = ttk.Entry(axis_frame, textvariable=y_label_var, width=30)
    y_label_entry.grid(row=1, column=1, padx=5, pady=5, sticky="w")

    ttk.Label(axis_frame, text="坐标轴标题字号:").grid(row=2, column=0, padx=5, pady=5, sticky="e")
    axis_title_fontsize_var = tk.StringVar(value=str(multi_temp_plot_settings['axis_title_fontsize']))
    axis_title_fontsize_entry = ttk.Entry(axis_frame, textvariable=axis_title_fontsize_var, width=10)
    axis_title_fontsize_entry.grid(row=2, column=1, padx=5, pady=5, sticky="w")

    ttk.Label(axis_frame, text="坐标轴刻度字号:").grid(row=3, column=0, padx=5, pady=5, sticky="e")
    axis_tick_fontsize_var = tk.StringVar(value=str(multi_temp_plot_settings['axis_tick_fontsize']))
    axis_tick_fontsize_entry = ttk.Entry(axis_frame, textvariable=axis_tick_fontsize_var, width=10)
    axis_tick_fontsize_entry.grid(row=3, column=1, padx=5, pady=5, sticky="w")

    # 图例设置
    show_legend_var = tk.BooleanVar(value=multi_temp_plot_settings['show_legend'])
    show_legend_cb = ttk.Checkbutton(legend_frame, text="显示图例", variable=show_legend_var)
    show_legend_cb.grid(row=0, column=0, padx=5, pady=5, sticky="w")

    # 新增：图例边框设置
    legend_frame_var = tk.BooleanVar(value=multi_temp_plot_settings['legend_frame'])
    legend_frame_cb = ttk.Checkbutton(legend_frame, text="图例边框", variable=legend_frame_var)
    legend_frame_cb.grid(row=0, column=1, padx=5, pady=5, sticky="w")

    # 标题设置
    show_title_var = tk.BooleanVar(value=multi_temp_plot_settings['show_title'])
    show_title_cb = ttk.Checkbutton(title_frame, text="显示标题", variable=show_title_var)
    show_title_cb.grid(row=0, column=0, padx=5, pady=5, sticky="w")

    ttk.Label(title_frame, text="标题文本:").grid(row=1, column=0, padx=5, pady=5, sticky="e")
    title_text_var = tk.StringVar(value=multi_temp_plot_settings['title_text'])
    title_text_entry = ttk.Entry(title_frame, textvariable=title_text_var, width=30)
    title_text_entry.grid(row=1, column=1, padx=5, pady=5, sticky="w")

    ttk.Label(title_frame, text="标题字号:").grid(row=2, column=0, padx=5, pady=5, sticky="e")
    title_fontsize_var = tk.StringVar(value=str(multi_temp_plot_settings['title_font_size']))
    title_fontsize_entry = ttk.Entry(title_frame, textvariable=title_fontsize_var, width=10)
    title_fontsize_entry.grid(row=2, column=1, padx=5, pady=5, sticky="w")

    # 其他选项
    show_grid_var = tk.BooleanVar(value=multi_temp_plot_settings['show_grid'])
    show_grid_cb = ttk.Checkbutton(options_frame, text="显示网格", variable=show_grid_var)
    show_grid_cb.grid(row=0, column=0, padx=5, pady=5, sticky="w")

    # 字体设置
    font_rows = [
        ("全局字体", "font_family", "font_size"),
        ("标题字体", "title_font_family", "title_font_size"),
        ("图例字体", "legend_font_family", "legend_font_size"),
        ("坐标轴标题字体", "axis_font_family", "axis_font_size"),
        ("刻度字体", "tick_font_family", "tick_font_size")
    ]

    # 存储字体标签和StringVar的字典
    font_labels = {}

    for i, (label, family_key, size_key) in enumerate(font_rows):
        ttk.Label(font_frame, text=f"{label}:").grid(row=i, column=0, padx=5, pady=5, sticky="e")

        # 显示当前字体设置
        current_family = multi_temp_plot_settings[family_key]
        current_size = multi_temp_plot_settings[size_key]
        current_font_label = f"{current_family}, {current_size}"
        font_var = tk.StringVar(value=current_font_label)
        font_label = ttk.Label(font_frame, textvariable=font_var)
        font_label.grid(row=i, column=1, padx=5, pady=5, sticky="w")

        # 存储标签和StringVar
        font_labels[(family_key, size_key)] = (font_var, font_label)

        # 字体选择按钮
        def create_choose_font(fk=family_key, sk=size_key, lbl=label):
            def choose_font():
                font_result = choose_font_dialog(
                    settings_window,
                    multi_temp_plot_settings[fk],
                    multi_temp_plot_settings[sk],
                    title=f"选择{lbl}字体"
                )
                if font_result:
                    multi_temp_plot_settings[fk] = font_result["family"]
                    multi_temp_plot_settings[sk] = font_result["size"]
                    # 更新显示
                    font_var, _ = font_labels[(fk, sk)]
                    font_var.set(f"{font_result['family']}, {font_result['size']}")

            return choose_font

        ttk.Button(font_frame, text="选择字体...",
                   command=create_choose_font()).grid(row=i, column=2, padx=5, pady=5)

    # 确认按钮
    def confirm_settings():
        try:
            multi_temp_plot_settings['line_width'] = float(line_width_var.get())
            multi_temp_plot_settings['axis_title_fontsize'] = int(axis_title_fontsize_var.get())
            multi_temp_plot_settings['axis_tick_fontsize'] = int(axis_tick_fontsize_var.get())
            multi_temp_plot_settings['show_legend'] = show_legend_var.get()
            multi_temp_plot_settings['legend_frame'] = legend_frame_var.get()  # 新增
            multi_temp_plot_settings['x_label'] = x_label_var.get()
            multi_temp_plot_settings['y_label'] = y_label_var.get()
            multi_temp_plot_settings['show_title'] = show_title_var.get()  # 新增
            multi_temp_plot_settings['title_text'] = title_text_var.get()  # 新增
            multi_temp_plot_settings['title_font_size'] = int(title_fontsize_var.get())  # 新增
            multi_temp_plot_settings['show_grid'] = show_grid_var.get()

            # 保存设置
            save_settings()
            settings_window.destroy()
        except ValueError as e:
            messagebox.showerror("输入错误", f"请输入有效的数值: {str(e)}")

    # 将确认按钮放在主框架底部，确保始终可见
    btn_frame = tk.Frame(main_frame)
    btn_frame.pack(side=tk.BOTTOM, fill=tk.X, padx=10, pady=10)

    ttk.Button(btn_frame, text="确认", command=confirm_settings).pack(pady=10)

    # 等待窗口关闭
    settings_window.wait_window()

    return multi_temp_plot_settings


def create_summary_file(folder_path, processed_files, include_c_point=True, preview_text=None):
    """创建汇总Excel文件，从所有处理过的文件中提取PotEng列数据，并进行指数拟合"""
    global a, b, c  # 使用全局变量

    try:
        folder_name = os.path.basename(folder_path)
        summary_file = os.path.join(folder_path, f"{folder_name}_summary.xlsx")

        if preview_text:
            preview_text.insert(tk.END, f"创建汇总文件: {os.path.basename(summary_file)}\n", "info")
            preview_text.see(tk.END)
            preview_text.update_idletasks()

        # 创建一个新的Excel writer对象
        with pd.ExcelWriter(summary_file, engine='openpyxl') as writer:
            summary_df = pd.DataFrame()
            source_info = []

            # 用于存储第一个文件的第一列数据
            first_col_data = None

            # 遍历所有处理过的文件
            for root, log_file, xlsx_file in processed_files:
                subfolder_name = os.path.basename(root)

                if preview_text:
                    preview_text.insert(tk.END, f"处理子文件夹: {subfolder_name}\n")
                    preview_text.see(tk.END)
                    preview_text.update_idletasks()

                try:
                    df = pd.read_excel(xlsx_file, header=None)

                    # 查找包含"PotEng"的列
                    poteng_col_idx = None
                    # 检查第一行是否有标题
                    for col_idx in range(df.shape[1]):
                        cell_value = str(df.iloc[0, col_idx]) if df.shape[0] > 0 else ""
                        if "PotEng" in cell_value:
                            poteng_col_idx = col_idx
                            break

                    # 如果没有找到PotEng列，尝试查找第二行（有些文件可能有两行标题）
                    if poteng_col_idx is None and df.shape[0] > 1:
                        for col_idx in range(df.shape[1]):
                            cell_value = str(df.iloc[1, col_idx])
                            if "PotEng" in cell_value:
                                poteng_col_idx = col_idx
                                break

                    # 如果仍然找不到PotEng列，尝试查找包含"Energy"或"Ener"的列
                    if poteng_col_idx is None:
                        for col_idx in range(df.shape[1]):
                            cell_value = str(df.iloc[0, col_idx]) if df.shape[0] > 0 else ""
                            if "Energy" in cell_value or "Ener" in cell_value or "energy" in cell_value.lower():
                                poteng_col_idx = col_idx
                                break

                    # 如果仍然找不到，使用第三列（通常PotEng在第三列）
                    if poteng_col_idx is None and df.shape[1] > 2:
                        poteng_col_idx = 2
                        if preview_text:
                            preview_text.insert(tk.END,
                                                f"警告: 在文件 {os.path.basename(xlsx_file)} 中未找到PotEng列，使用第三列作为替代\n",
                                                "warning")
                            preview_text.see(tk.END)
                            preview_text.update_idletasks()

                    if poteng_col_idx is None:
                        if preview_text:
                            preview_text.insert(tk.END,
                                                f"警告: 在文件 {os.path.basename(xlsx_file)} 中未找到PotEng列，跳过\n",
                                                "warning")
                            preview_text.see(tk.END)
                            preview_text.update_idletasks()
                        continue

                    # 提取PotEng列数据（从第二行开始）
                    poteng_data = df.iloc[1:, poteng_col_idx].reset_index(drop=True)

                    # 保存第一个文件的第一列数据（A列，从第二行开始）
                    if first_col_data is None and df.shape[1] > 0:
                        first_col_data = df.iloc[1:, 0].reset_index(drop=True)

                    # 确保列名唯一
                    col_name = subfolder_name
                    suffix = 1
                    while col_name in summary_df.columns:
                        col_name = f"{subfolder_name}_{suffix}"
                        suffix += 1

                    if summary_df.empty:
                        summary_df = pd.DataFrame({col_name: poteng_data})
                    else:
                        # 确保行数一致
                        if len(poteng_data) > len(summary_df):
                            # 扩展现有DataFrame
                            diff = len(poteng_data) - len(summary_df)
                            summary_df = summary_df.reindex(range(len(summary_df) + diff))
                        elif len(poteng_data) < len(summary_df):
                            # 扩展新列数据
                            diff = len(summary_df) - len(poteng_data)
                            poteng_data = pd.concat([poteng_data, pd.Series([""] * diff)], ignore_index=True)

                        summary_df[col_name] = poteng_data

                    # 记录来源信息
                    source_info.append({
                        '子文件夹': subfolder_name,
                        '日志文件': os.path.basename(log_file),
                        '数据文件': os.path.basename(xlsx_file),
                        'PotEng列索引': poteng_col_idx
                    })

                except Exception as e:
                    if preview_text:
                        preview_text.insert(tk.END, f"提取文件 {xlsx_file} 数据失败: {str(e)}\n", "error")
                        preview_text.see(tk.END)
                        preview_text.update_idletasks()

            # 添加第一列数据（Step列）
            if not summary_df.empty:
                # 添加Step列作为第一列
                if first_col_data is not None:
                    # 调整长度以匹配汇总数据
                    if len(first_col_data) > len(summary_df):
                        first_col_data = first_col_data.iloc[:len(summary_df)]
                    elif len(first_col_data) < len(summary_df):
                        diff = len(summary_df) - len(first_col_data)
                        first_col_data = pd.concat([first_col_data, pd.Series([""] * diff)], ignore_index=True)

                    summary_df.insert(0, "Step", first_col_data)
                else:
                    # 如果没有Step数据，创建一个空白列
                    blank_col = pd.Series([""] * len(summary_df))
                    summary_df.insert(0, "Step", blank_col)

            # 添加Time/ps列
            if not summary_df.empty and summary_df.shape[0] >= 2:
                # 创建新的"Time/ps"列
                time_ps_col = pd.Series([""] * len(summary_df))

                # 从第二行开始计算值（索引1），并处理0值
                for i in range(1, len(summary_df)):
                    try:
                        # 获取Step列的值
                        step_value = summary_df.iloc[i, 0]

                        # 直接处理0值情况
                        if step_value == 0 or step_value == "0":
                            time_ps_col.iloc[i] = 0.0
                        else:
                            # 转换为数值并除以10000
                            time_ps = float(step_value) / 10000.0
                            time_ps_col.iloc[i] = time_ps
                    except (ValueError, TypeError):
                        # 如果转换失败，保留空值
                        time_ps_col.iloc[i] = ""

                # 在第二列位置插入"Time/ps"列
                summary_df.insert(1, "Time/ps", time_ps_col)

            # 将汇总数据保存到Excel
            if not summary_df.empty:
                summary_df.to_excel(writer, index=False, sheet_name="数据汇总")
                if preview_text:
                    preview_text.insert(tk.END, "已创建'数据汇总'工作表\n")
                    preview_text.see(tk.END)
                    preview_text.update_idletasks()

            # 添加来源信息
            if source_info:
                source_df = pd.DataFrame(source_info)
                source_df.to_excel(writer, index=False, sheet_name="来源信息")
                if preview_text:
                    preview_text.insert(tk.END, "已创建'来源信息'工作表\n")
                    preview_text.see(tk.END)
                    preview_text.update_idletasks()

            # 自动调整列宽
            workbook = writer.book
            for sheetname in workbook.sheetnames:
                worksheet = workbook[sheetname]
                for col in worksheet.columns:
                    max_length = 0
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = max_length + 2
                    worksheet.column_dimensions[col[0].column_letter].width = adjusted_width

        # ================== 新增功能：为每个温度创建单独的工作表并进行指数拟合 ==================
        wb = load_workbook(summary_file)
        if "数据汇总" in wb.sheetnames:
            ws_summary = wb["数据汇总"]

            if preview_text:
                preview_text.insert(tk.END, "开始分析数据...\n")
                preview_text.see(tk.END)
                preview_text.update_idletasks()

            # 创建一个新的工作表来存储所有温度下的k值
            k_sheet_name = "反应速率常数"
            if k_sheet_name in wb.sheetnames:
                wb.remove(wb[k_sheet_name])
            ws_k = wb.create_sheet(title=k_sheet_name)

            if preview_text:
                preview_text.insert(tk.END, f"已创建'{k_sheet_name}'工作表\n")
                preview_text.see(tk.END)
                preview_text.update_idletasks()

            # 在k值汇总表中添加是否包含X=c点的信息
            ws_k.append(["是否包含X=c的数据点", "是" if include_c_point else "否"])
            ws_k.append(["温度 (K)", "反应速率常数k", "拟合优度R²", "拟合数据点数", "拟合公式"])

            # 获取数据区域的最大行和最大列
            max_row = ws_summary.max_row
            max_col = ws_summary.max_column

            # 从第三列开始处理每一列温度数据（跳过Step列和Time/ps列）
            for col_idx in range(3, max_col + 1):
                # 获取温度值作为工作表名称
                temp_value = ws_summary.cell(row=1, column=col_idx).value
                if temp_value is None:
                    continue

                if preview_text:
                    preview_text.insert(tk.END, f"分析温度: {temp_value} K\n")
                    preview_text.see(tk.END)
                    preview_text.update_idletasks()

                # 创建新的工作表
                try:
                    # 工作表名称不能包含特殊字符，截断到31个字符
                    sheet_name = str(temp_value)[:30]
                    if sheet_name in wb.sheetnames:
                        ws_temp = wb[sheet_name]
                    else:
                        ws_temp = wb.create_sheet(title=sheet_name)
                except:
                    # 如果名称无效，使用默认名称
                    sheet_name = f"Temperature_{col_idx - 2}"
                    ws_temp = wb.create_sheet(title=sheet_name)

                if preview_text:
                    preview_text.insert(tk.END, f"已创建工作表: {sheet_name}\n")
                    preview_text.see(tk.END)
                    preview_text.update_idletasks()

                # 添加标题行
                if ws_temp.max_row == 1:
                    ws_temp.append(["Time/ps", "Energy"])

                # 收集时间和能量数据
                time_values = []
                energy_values = []

                # 从第二行开始读取数据（第一行是标题）
                for row_idx in range(2, max_row + 1):
                    # 获取时间值（第二列）
                    time_cell = ws_summary.cell(row=row_idx, column=2).value
                    # 获取能量值（当前温度列）
                    energy_cell = ws_summary.cell(row=row_idx, column=col_idx).value

                    # 只添加有效数据
                    if time_cell is not None and energy_cell is not None:
                        try:
                            # 尝试转换为数值
                            time_val = float(time_cell)
                            energy_val = float(energy_cell)
                            time_values.append(time_val)
                            energy_values.append(energy_val)
                        except (ValueError, TypeError):
                            # 转换失败则跳过
                            continue

                # 将数据写入新工作表（不再跳过第二行）
                for i in range(len(time_values)):
                    if i >= ws_temp.max_row - 1:  # 避免覆盖已有的统计结果
                        ws_temp.append([time_values[i], energy_values[i]])

                # 计算统计值
                if len(energy_values) > 0:
                    # 1. 计算最大值及所在时间
                    max_val = max(energy_values)
                    max_index = energy_values.index(max_val)
                    max_time = time_values[max_index] if max_index < len(time_values) else "N/A"

                    # 2. 计算最后10个数据的平均值
                    last_10 = energy_values[-10:]
                    avg_val = sum(last_10) / len(last_10) if last_10 else 0

                    # 3. 计算最大值与平均值的差值
                    diff_val = max_val - avg_val

                    # 将计算结果赋值给全局变量
                    a = avg_val
                    b = diff_val
                    c = max_time

                    # 在工作表末尾添加统计结果
                    ws_temp.append([])  # 空行
                    ws_temp.append(["统计结果", ""])
                    ws_temp.append(["最大值", max_val])
                    ws_temp.append(["最大值所在时间", max_time])
                    ws_temp.append(["最后10个平均值", avg_val])
                    ws_temp.append(["差值(最大值-平均值)", diff_val])

                    # 添加说明
                    ws_temp.append([])
                    ws_temp.append(["说明", ""])
                    ws_temp.append(["c = 最大值所在时间", c])
                    ws_temp.append(["a = 最后10个平均值", a])
                    ws_temp.append(["b = 差值(最大值-平均值)", b])
                    ws_temp.append(["是否包含X=c的数据点", "是" if include_c_point else "否"])

                    # ================== 新增：指数拟合 ==================
                    # 使用公式 Y = a + b * exp[-k*(X - c)] 进行拟合
                    # 准备用于拟合的数据
                    fit_x = []
                    fit_y = []

                    # 收集所有X>=c的数据点（包括X=c）
                    for i in range(len(time_values)):
                        # 根据用户选择决定是否包含X=c的点
                        if include_c_point:
                            condition = (time_values[i] >= c) and (energy_values[i] > a)
                        else:
                            condition = (time_values[i] > c) and (energy_values[i] > a)

                        if condition:
                            fit_x.append(time_values[i])
                            fit_y.append(energy_values[i])

                    # 检查是否有足够的数据点进行拟合
                    if len(fit_x) >= 3:  # 需要至少3个点进行拟合
                        try:
                            if preview_text:
                                preview_text.insert(tk.END, f"进行指数拟合，数据点: {len(fit_x)}个\n")
                                preview_text.see(tk.END)
                                preview_text.update_idletasks()

                            # 使用curve_fit进行非线性拟合
                            popt, pcov = curve_fit(exponential_func, fit_x, fit_y, p0=[0.01], bounds=(0, np.inf))

                            # 提取拟合参数
                            k = popt[0]

                            # 计算拟合优度 R²
                            residuals = fit_y - exponential_func(np.array(fit_x), k)
                            ss_res = np.sum(residuals ** 2)
                            ss_tot = np.sum((np.array(fit_y) - np.mean(fit_y)) ** 2)
                            r_squared = 1 - (ss_res / ss_tot) if ss_tot != 0 else 0

                            # 添加拟合结果
                            ws_temp.append([])
                            ws_temp.append(["指数拟合结果"])
                            ws_temp.append(["反应速率常数k", k])
                            ws_temp.append(["拟合优度R²", r_squared])
                            ws_temp.append(["拟合数据点数", len(fit_x)])
                            ws_temp.append(["拟合公式", f"Y = {a:.4f} + {b:.4f} * exp[-{k:.4f}*(X - {c:.4f})]"])

                            # 添加到k值汇总表
                            ws_k.append([temp_value, k, r_squared, len(fit_x),
                                         f"Y = {a:.4f} + {b:.4f} * exp[-{k:.4f}*(X - {c:.4f})]"])

                            if preview_text:
                                preview_text.insert(tk.END, f"拟合成功: k = {k:.4f}, R² = {r_squared:.4f}\n", "success")
                                preview_text.see(tk.END)
                                preview_text.update_idletasks()

                        except Exception as e:
                            error_msg = f"指数拟合失败: {str(e)}"
                            if preview_text:
                                preview_text.insert(tk.END, f"错误: {error_msg}\n", "error")
                                preview_text.see(tk.END)
                                preview_text.update_idletasks()

                            ws_temp.append([])
                            ws_temp.append(["指数拟合结果"])
                            ws_temp.append(["错误: 无法完成指数拟合", str(e)])
                    else:
                        error_msg = f"有效数据点不足 (需要至少3个X>=c的数据点)，实际: {len(fit_x)}"
                        if preview_text:
                            preview_text.insert(tk.END, f"错误: {error_msg}\n", "error")
                            preview_text.see(tk.END)
                            preview_text.update_idletasks()

                        ws_temp.append([])
                        ws_temp.append(["指数拟合结果"])
                        ws_temp.append(["错误: 有效数据点不足 (需要至少3个X>=c的数据点)"])
                        ws_temp.append(["有效数据点数", len(fit_x)])

            # ================== 新增：活化能拟合 ==================
            # 检查是否有足够的k值数据
            if ws_k.max_row > 2:  # 至少有一行数据（标题行+设置行+至少一行数据）
                if preview_text:
                    preview_text.insert(tk.END, "开始活化能拟合...\n")
                    preview_text.see(tk.END)
                    preview_text.update_idletasks()

                # 准备数据用于活化能拟合
                temperatures = []
                k_values = []

                # 从第三行开始读取数据（第一行是设置，第二行是标题）
                for row in range(3, ws_k.max_row + 1):
                    temp_val = ws_k.cell(row=row, column=1).value
                    k_val = ws_k.cell(row=row, column=2).value

                    if temp_val is not None and k_val is not None:
                        try:
                            # 处理温度值（去除可能的后缀并转换为数值）
                            if isinstance(temp_val, str):
                                # 去除可能的'T'后缀
                                if temp_val.endswith('T'):
                                    temp_val = temp_val.rstrip('T')
                                # 去除可能的'K'后缀
                                if temp_val.endswith('K'):
                                    temp_val = temp_val.rstrip('K')
                            temperature = float(temp_val)
                            k_value = float(k_val)

                            temperatures.append(temperature)
                            k_values.append(k_value)
                        except (ValueError, TypeError):
                            continue

                # 检查是否有足够的数据点进行线性拟合
                if len(temperatures) >= 3:
                    try:
                        # 准备数据：X = 1000/T (K^{-1}), Y = ln(k)
                        inv_T = 1000.0 / np.array(temperatures)  # 1000/T
                        ln_k = np.log(k_values)

                        # 进行线性拟合 (使用公式 lnk = lnA - (Ea/R) * (1000/T))
                        slope, intercept, r_value, p_value, std_err = linregress(inv_T, ln_k)

                        # 计算活化能和指前因子
                        # 根据公式 lnk = lnA - (Ea/R) * (1000/T)
                        # 所以斜率 = -Ea/R, 截距 = lnA
                        Ea = -slope * R  # 活化能 (J/mol)
                        lnA = intercept
                        A = np.exp(lnA)  # 指前因子

                        # 创建新的工作表
                        activation_sheet = wb.create_sheet(title="活化能拟合")

                        if preview_text:
                            preview_text.insert(tk.END, "已创建'活化能拟合'工作表\n")
                            preview_text.see(tk.END)
                            preview_text.update_idletasks()

                        # 写入标题
                        activation_sheet.append(["温度 (K)", "1000/T (K⁻¹)", "k", "ln(k)"])

                        # 写入数据
                        for i in range(len(temperatures)):
                            activation_sheet.append([
                                temperatures[i],
                                inv_T[i],
                                k_values[i],
                                ln_k[i]
                            ])

                        # 添加空行
                        activation_sheet.append([])
                        activation_sheet.append([])

                        # 添加拟合结果
                        activation_sheet.append(["拟合结果", "", "", ""])
                        activation_sheet.append(["斜率 (slope)", slope])
                        activation_sheet.append(["截距 (intercept)", intercept])
                        activation_sheet.append(["相关系数 (R)", r_value])
                        activation_sheet.append(["R²", r_value ** 2])
                        activation_sheet.append(["标准误差", std_err])
                        activation_sheet.append(["p值", p_value])
                        activation_sheet.append([])
                        activation_sheet.append(["活化能 Ea (J/mol)", Ea])
                        activation_sheet.append(["活化能 Ea (kJ/mol)", Ea / 1000])
                        activation_sheet.append(["指前因子 A", A])
                        activation_sheet.append(["ln(A)", lnA])
                        activation_sheet.append([])
                        activation_sheet.append(["拟合公式", f"ln(k) = {intercept:.4f} + ({slope:.4f}) * (1000/T)"])
                        activation_sheet.append(["", f"ln(k) = ln(A) - (Ea/R) * (1000/T)"])

                        if preview_text:
                            preview_text.insert(tk.END,
                                                f"活化能拟合成功: Ea = {Ea / 1000:.2f} kJ/mol, R² = {r_value ** 2:.4f}\n",
                                                "success")
                            preview_text.see(tk.END)
                            preview_text.update_idletasks()

                        # ================== 新增：生成活化能拟合图 ==================
                        # 获取用户设置
                        settings = get_activation_plot_settings()

                        # 设置字体 - 使用正确的参数名
                        plt.rcParams['font.family'] = settings['font_family']
                        plt.rcParams['font.size'] = settings['font_size']
                        plt.rcParams['legend.fontsize'] = settings['legend_font_size']
                        plt.rcParams['axes.titlesize'] = settings['axis_font_size']
                        plt.rcParams['axes.labelsize'] = settings['axis_font_size']
                        plt.rcParams['xtick.labelsize'] = settings['tick_font_size']
                        plt.rcParams['ytick.labelsize'] = settings['tick_font_size']

                        # 创建图形
                        fig, ax = plt.subplots(figsize=(8, 6))

                        # 绘制原始数据点
                        ax.scatter(
                            inv_T,
                            ln_k,
                            s=settings['scatter_size'],
                            color=settings['scatter_color'],
                            label=settings['legend_label_data']
                        )

                        # 绘制拟合直线
                        # 创建拟合线的x值范围（从最小值到最大值）
                        x_fit = np.linspace(min(inv_T), max(inv_T), 100)
                        y_fit = intercept + slope * x_fit
                        ax.plot(
                            x_fit,
                            y_fit,
                            color=settings['line_color'],
                            linewidth=settings['line_width'],
                            label=settings['legend_label_fit']
                        )

                        # 设置坐标轴标签
                        ax.set_xlabel(settings['x_label'], fontsize=settings['axis_title_fontsize'])
                        ax.set_ylabel(settings['y_label'], fontsize=settings['axis_title_fontsize'])

                        # 设置刻度字体大小
                        ax.tick_params(axis='both', which='major', labelsize=settings['axis_tick_fontsize'])

                        # 添加网格
                        if settings['show_grid']:
                            ax.grid(True, linestyle='--', alpha=0.7)

                        # 添加图例
                        if settings['show_legend']:
                            legend = ax.legend(fontsize=settings['legend_font_size'], loc='best')
                            # 设置图例边框
                            legend.get_frame().set_visible(settings['legend_frame'])

                        # 添加统计信息
                        if settings['show_statistics']:
                            text_str = f"ln(k) = {intercept:.4f} + {slope:.4f}·(1000/T)\nR² = {r_value ** 2:.4f}\nEa = {Ea / 1000:.2f} kJ/mol"
                            ax.text(
                                settings['statistics_position_x'],
                                settings['statistics_position_y'],
                                text_str,
                                transform=ax.transAxes,
                                fontsize=settings['statistics_fontsize'],
                                verticalalignment='top',
                                bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.5)
                            )

                        # 保存图像到临时文件
                        img_path = os.path.join(folder_path, "activation_energy_plot.png")
                        fig.savefig(img_path, dpi=300, bbox_inches='tight')
                        plt.close(fig)

                        if preview_text:
                            preview_text.insert(tk.END, "已生成活化能拟合图\n")
                            preview_text.see(tk.END)
                            preview_text.update_idletasks()

                        # 将图像插入到Excel
                        img = openpyxl.drawing.image.Image(img_path)
                        img.anchor = f'A{len(temperatures) + 15}'  # 从数据行下面15行开始插入
                        activation_sheet.add_image(img)

                    except Exception as e:
                        error_msg = f"活化能拟合失败: {str(e)}"
                        if preview_text:
                            preview_text.insert(tk.END, f"错误: {error_msg}\n", "error")
                            preview_text.see(tk.END)
                            preview_text.update_idletasks()

                        # 创建新的工作表并写入错误信息
                        activation_sheet = wb.create_sheet(title="活化能拟合")
                        activation_sheet.append(["活化能拟合失败", str(e)])
                else:
                    error_msg = f"有效数据点不足 (需要至少3个温度点的k值)，实际: {len(temperatures)}"
                    if preview_text:
                        preview_text.insert(tk.END, f"错误: {error_msg}\n", "error")
                        preview_text.see(tk.END)
                        preview_text.update_idletasks()

                    # 创建新的工作表并写入错误信息
                    activation_sheet = wb.create_sheet(title="活化能拟合")
                    activation_sheet.append(["错误: 有效数据点不足 (需要至少3个温度点的k值)"])
                    activation_sheet.append(["有效温度点数", len(temperatures)])

            # ================== 新增：生成多温度能量-时间变化图 ==================
            # 检查是否有足够的数据
            if "数据汇总" in wb.sheetnames and ws_summary.max_row > 1 and ws_summary.max_column >= 3:
                try:
                    # 创建新的工作表
                    multi_temp_sheet = wb.create_sheet(title="能量-时间变化图")

                    if preview_text:
                        preview_text.insert(tk.END, "已创建'能量-时间变化图'工作表\n")
                        preview_text.see(tk.END)
                        preview_text.update_idletasks()

                    # 获取绘图设置
                    settings = get_multi_temp_plot_settings()

                    # 设置字体 - 使用正确的参数名
                    plt.rcParams['font.family'] = settings['font_family']
                    plt.rcParams['font.size'] = settings['font_size']
                    plt.rcParams['legend.fontsize'] = settings['legend_font_size']
                    plt.rcParams['axes.titlesize'] = settings['axis_font_size']
                    plt.rcParams['axes.labelsize'] = settings['axis_font_size']
                    plt.rcParams['xtick.labelsize'] = settings['tick_font_size']
                    plt.rcParams['ytick.labelsize'] = settings['tick_font_size']

                    # 创建图形
                    fig, ax = plt.subplots(figsize=(10, 7))

                    # 获取时间数据（第二列）
                    time_data = []
                    for row_idx in range(2, ws_summary.max_row + 1):
                        time_val = ws_summary.cell(row=row_idx, column=2).value
                        if time_val is not None:
                            try:
                                time_data.append(float(time_val))
                            except:
                                pass

                    # 如果没有时间数据，使用行号作为X轴
                    if not time_data:
                        time_data = list(range(1, ws_summary.max_row))

                    # 为每个温度绘制曲线
                    for col_idx in range(3, ws_summary.max_column + 1):
                        # 获取温度值
                        temp_val = ws_summary.cell(row=1, column=col_idx).value
                        if temp_val is None:
                            continue

                        # 获取能量数据
                        energy_data = []
                        for row_idx in range(2, ws_summary.max_row + 1):
                            energy_val = ws_summary.cell(row=row_idx, column=col_idx).value
                            if energy_val is not None:
                                try:
                                    energy_data.append(float(energy_val))
                                except:
                                    pass

                        # 确保数据长度一致
                        min_len = min(len(time_data), len(energy_data))
                        if min_len > 0:
                            # 绘制曲线
                            ax.plot(
                                time_data[:min_len],
                                energy_data[:min_len],
                                label=f"{temp_val} K",
                                linewidth=settings['line_width']
                            )

                    # 设置坐标轴标签
                    ax.set_xlabel(settings['x_label'], fontsize=settings['axis_title_fontsize'])
                    ax.set_ylabel(settings['y_label'], fontsize=settings['axis_title_fontsize'])

                    # 设置刻度字体大小
                    ax.tick_params(axis='both', which='major', labelsize=settings['axis_tick_fontsize'])

                    # 添加网格
                    if settings['show_grid']:
                        ax.grid(True, linestyle='--', alpha=0.7)

                    # 添加图例
                    if settings['show_legend']:
                        legend = ax.legend(fontsize=settings['legend_font_size'], loc='best')
                        # 设置图例边框
                        legend.get_frame().set_visible(settings['legend_frame'])

                    # 添加标题
                    if settings['show_title']:  # 新增：是否显示标题
                        ax.set_title(settings['title_text'],
                                     fontsize=settings['title_font_size'])

                    # 保存图像到临时文件
                    img_path = os.path.join(folder_path, "multi_temp_plot.png")
                    fig.savefig(img_path, dpi=300, bbox_inches='tight')
                    plt.close(fig)

                    if preview_text:
                        preview_text.insert(tk.END, "已生成多温度能量图\n")
                        preview_text.see(tk.END)
                        preview_text.update_idletasks()

                    # 将图像插入到Excel
                    img = openpyxl.drawing.image.Image(img_path)
                    multi_temp_sheet.add_image(img, "A1")

                except Exception as e:
                    error_msg = f"生成多温度能量图失败: {str(e)}"
                    if preview_text:
                        preview_text.insert(tk.END, f"错误: {error_msg}\n", "error")
                        preview_text.see(tk.END)
                        preview_text.update_idletasks()

                    # 创建新的工作表并写入错误信息
                    multi_temp_sheet = wb.create_sheet(title="能量-时间变化图")
                    multi_temp_sheet.append(["生成多温度能量图失败", str(e)])

            # 重新调整所有工作表的列宽
            for sheet in wb:
                for col in sheet.columns:
                    max_length = 0
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = max_length + 2
                    sheet.column_dimensions[col[0].column_letter].width = adjusted_width

            # 保存修改后的工作簿
            wb.save(summary_file)
            if preview_text:
                preview_text.insert(tk.END, "已保存汇总文件\n\n", "success")
                preview_text.see(tk.END)
                preview_text.update_idletasks()
        # ================== 新增功能结束 ==================

        return summary_file

    except Exception as e:
        error_msg = f"创建汇总文件失败: {str(e)}"
        if preview_text:
            preview_text.insert(tk.END, f"错误: {error_msg}\n", "error")
            preview_text.see(tk.END)
            preview_text.update_idletasks()
        raise RuntimeError(error_msg) from e


def preview_plot(image_path, title="预览图"):
    """在单独窗口中预览图像"""
    preview_window = tk.Toplevel()
    preview_window.title(title)
    preview_window.geometry("800x600")

    # 加载图像
    try:
        img = Image.open(image_path)
        img.thumbnail((780, 580))  # 调整大小以适应窗口

        # 转换为Tkinter格式
        img_tk = ImageTk.PhotoImage(img)

        # 创建标签显示图像
        label = tk.Label(preview_window, image=img_tk)
        label.image = img_tk  # 保持引用
        label.pack(padx=10, pady=10)

        # 添加关闭按钮
        btn_close = ttk.Button(preview_window, text="关闭", command=preview_window.destroy)
        btn_close.pack(pady=10)
    except Exception as e:
        tk.Label(preview_window, text=f"无法加载图像: {str(e)}").pack(padx=10, pady=10)


def main():
    """主函数，处理文件夹选择和处理过程"""
    global include_c_point  # 使用全局变量

    root = tk.Tk()
    root.withdraw()
    root.title("日志文件批量处理工具")

    folder_path = filedialog.askdirectory(title="选择包含日志文件的文件夹")
    if not folder_path:
        messagebox.showinfo("信息", "操作已取消")
        return

    # 创建主预览窗口
    preview_window = tk.Toplevel(root)
    preview_window.title("处理进度预览")
    preview_window.geometry("900x700")

    # 创建标签框架
    frame = ttk.Frame(preview_window)
    frame.pack(fill="both", expand=True, padx=10, pady=10)

    # 创建进度标签
    tk.Label(frame, text="处理进度:", font=("Arial", 12, "bold")).pack(anchor="w", pady=(0, 5))

    # 创建带滚动条的文本区域
    preview_text = scrolledtext.ScrolledText(frame, wrap=tk.WORD, width=100, height=20)
    preview_text.pack(fill="both", expand=True, pady=(0, 10))

    # 配置标签样式
    preview_text.tag_config("info", foreground="blue")
    preview_text.tag_config("success", foreground="green")
    preview_text.tag_config("warning", foreground="orange")
    preview_text.tag_config("error", foreground="red")

    # 添加分隔线
    ttk.Separator(frame).pack(fill="x", pady=10)

    # 创建图像预览区域
    tk.Label(frame, text="图表预览:", font=("Arial", 12, "bold")).pack(anchor="w", pady=(0, 5))

    # 创建图像框架
    img_frame = ttk.Frame(frame)
    img_frame.pack(fill="both", expand=True)

    # 创建图像占位符
    img_label1 = tk.Label(img_frame, text="活化能拟合图预览", relief="groove", width=30, height=15)
    img_label1.grid(row=0, column=0, padx=10, pady=10)

    img_label2 = tk.Label(img_frame, text="多温度能量图预览", relief="groove", width=30, height=15)
    img_label2.grid(row=0, column=1, padx=10, pady=10)

    # 添加预览按钮
    btn_frame = ttk.Frame(img_frame)
    btn_frame.grid(row=1, column=0, columnspan=2, pady=10)

    btn_plot1 = ttk.Button(btn_frame, text="查看活化能拟合图", command=lambda: preview_plot(
        os.path.join(folder_path, "activation_energy_plot.png"), "活化能拟合图预览"))
    btn_plot1.grid(row=0, column=0, padx=10)

    btn_plot2 = ttk.Button(btn_frame, text="查看多温度能量图", command=lambda: preview_plot(
        os.path.join(folder_path, "multi_temp_plot.png"), "多温度能量图预览"))
    btn_plot2.grid(row=0, column=1, padx=10)

    # 询问用户是否包含X=c的数据点
    include_c_point = messagebox.askyesno(
        "选择拟合数据点",
        "是否在指数拟合中包含X=c的数据点?\n\n"
        "选择'是'将包含X=c的点\n"
        "选择'否'将只包含X>c的点"
    )

    try:
        # 更新预览窗口
        preview_text.insert(tk.END, "开始处理文件...\n\n")
        preview_text.see(tk.END)
        preview_text.update_idletasks()

        # 查找所有日志文件
        preview_text.insert(tk.END, "搜索日志文件中...\n", "info")
        preview_text.see(tk.END)
        preview_text.update_idletasks()

        log_files = find_log_files(folder_path, preview_text)
        total_files = len(log_files)

        if total_files == 0:
            preview_text.insert(tk.END, "未找到任何日志文件\n", "warning")
            preview_text.see(tk.END)
            preview_text.update_idletasks()
            messagebox.showinfo("信息", "未找到任何日志文件")
            return

        preview_text.insert(tk.END, f"共找到 {total_files} 个日志文件，开始处理...\n\n", "info")
        preview_text.see(tk.END)
        preview_text.update_idletasks()

        # 处理文件
        processed_files = []
        processed_count = 0

        for i, log_file in enumerate(log_files):
            try:
                preview_text.insert(tk.END, f"处理文件中: {os.path.basename(log_file)}\n")
                preview_text.see(tk.END)
                preview_text.update_idletasks()

                # 处理日志文件
                xlsx_file = process_log_file(log_file, preview_text)
                processed_files.append((os.path.dirname(log_file), log_file, xlsx_file))
                processed_count += 1

            except Exception as e:
                preview_text.insert(tk.END, f"处理失败: {str(e)}\n\n", "error")
                preview_text.see(tk.END)
                preview_text.update_idletasks()

        # 创建汇总文件
        if processed_count > 0:
            preview_text.insert(tk.END, f"\n成功处理 {processed_count} 个文件，开始创建汇总文件...\n\n", "info")
            preview_text.see(tk.END)
            preview_text.update_idletasks()

            # 将用户选择传递给汇总函数
            summary_file = create_summary_file(folder_path, processed_files, include_c_point, preview_text)

            # 显示结果
            c_point_status = "包含" if include_c_point else "不包含"
            result_msg = (
                f"处理完成！\n\n"
                f"共找到 {total_files} 个日志文件\n"
                f"成功处理 {processed_count} 个文件\n"
                f"汇总文件已保存到: {summary_file}\n\n"
                f"拟合设置: {c_point_status} X=c的数据点\n"
                f"每个温度的数据已提取到单独的工作表\n"
                f"已进行活化能拟合分析并生成图表"
            )

            preview_text.insert(tk.END, "\n" + "-" * 80 + "\n", "info")
            preview_text.insert(tk.END, "处理完成！\n\n", "success")
            preview_text.insert(tk.END, result_msg + "\n\n", "info")
            preview_text.see(tk.END)
            preview_text.update_idletasks()

            # 尝试加载并显示预览图像
            try:
                # 活化能拟合图
                activation_img = Image.open(os.path.join(folder_path, "activation_energy_plot.png"))
                activation_img.thumbnail((300, 250))
                activation_img_tk = ImageTk.PhotoImage(activation_img)
                img_label1.config(image=activation_img_tk)
                img_label1.image = activation_img_tk

                # 多温度能量图
                multi_temp_img = Image.open(os.path.join(folder_path, "multi_temp_plot.png"))
                multi_temp_img.thumbnail((300, 250))
                multi_temp_img_tk = ImageTk.PhotoImage(multi_temp_img)
                img_label2.config(image=multi_temp_img_tk)
                img_label2.image = multi_temp_img_tk
            except:
                pass

            # 添加操作按钮
            btn_frame = ttk.Frame(frame)
            btn_frame.pack(fill="x", pady=10)

            def open_summary():
                try:
                    if sys.platform == "win32":
                        os.startfile(summary_file)
                    elif sys.platform == "darwin":
                        import subprocess
                        subprocess.call(["open", summary_file])
                    else:
                        import subprocess
                        subprocess.call(["xdg-open", summary_file])
                except:
                    messagebox.showinfo("提示", "无法自动打开文件，请手动访问。")

            def open_folder():
                try:
                    if sys.platform == "win32":
                        os.startfile(folder_path)
                    elif sys.platform == "darwin":
                        import subprocess
                        subprocess.call(["open", folder_path])
                    else:
                        import subprocess
                        subprocess.call(["xdg-open", folder_path])
                except:
                    messagebox.showinfo("提示", "无法打开文件夹，请手动访问。")

            def close_app():
                root.destroy()

            ttk.Button(btn_frame, text="打开汇总文件", command=open_summary, width=15).pack(side=tk.LEFT, padx=10)
            ttk.Button(btn_frame, text="打开文件夹", command=open_folder, width=15).pack(side=tk.LEFT, padx=10)
            ttk.Button(btn_frame, text="退出", command=close_app, width=15).pack(side=tk.LEFT, padx=10)
        else:
            preview_text.insert(tk.END, "\n没有成功处理任何文件\n", "warning")
            preview_text.see(tk.END)
            preview_text.update_idletasks()
            messagebox.showinfo("信息", "没有成功处理任何文件")

    except Exception as e:
        preview_text.insert(tk.END, f"\n处理失败: {str(e)}\n", "error")
        preview_text.see(tk.END)
        preview_text.update_idletasks()
        messagebox.showerror("处理失败", str(e))

    # 禁用主窗口关闭按钮
    preview_window.protocol("WM_DELETE_WINDOW", lambda: None)

    # 等待预览窗口关闭
    preview_window.wait_window()
    root.destroy()


if __name__ == "__main__":
    main()