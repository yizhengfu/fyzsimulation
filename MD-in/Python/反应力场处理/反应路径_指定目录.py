import openpyxl

# 读取原始文件
input_file = r'E:\onechainGAP\reac.xlsx'
workbook = openpyxl.load_workbook(input_file)
sheet = workbook.active

# 创建新的工作簿
output_file = r'E:\onechainGAP\反应路径.xlsx'
output_workbook = openpyxl.Workbook()
output_sheet = output_workbook.active

# 初始化新表格的标题行
output_sheet['A1'] = 'Timestep'
output_sheet['B1'] = 'Reaction'

row_index = 2  # 从第二行开始写入新表格
for row in sheet.iter_rows(min_row=6):
    data_cells = [cell.value for cell in row[1:] if cell.value is not None and cell.value != 0]

    if data_cells:
        for cell in row[1:]:
            if cell.value is not None:
                if cell.value > 0:
                    output_sheet[f'A{row_index}'] = row[0].value
                    output_sheet[f'B{row_index}'] = sheet.cell(row=3, column=cell.column).value
                    row_index += 1
                elif cell.value < 0:
                    output_sheet[f'A{row_index}'] = row[0].value
                    original_text = sheet.cell(row=3, column=cell.column).value
                    reversed_text = original_text.split(' -> ')[1] + ' -> ' + original_text.split(' -> ')[0]
                    output_sheet[f'B{row_index}'] = reversed_text
                    row_index += 1

# 保存新表格
output_workbook.save(output_file)
