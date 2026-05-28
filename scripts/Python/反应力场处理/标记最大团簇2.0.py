import openpyxl

# 打开Excel文件
wb = openpyxl.load_workbook('2.xlsx')

# 选择要操作的工作表，这里假设你要处理的工作表名称为Sheet1
sheet = wb['Sheet1']

# 获取工作表的行数和列数
max_row = sheet.max_row
max_col = sheet.max_column

# 创建一个新的工作簿
new_wb = openpyxl.Workbook()
new_sheet = new_wb.active

# 循环遍历偶数行，从第2行开始，步进为2
for row_num in range(2, max_row + 1, 2):
    max_value = None
    max_col_letter = None

    # 遍历当前行的第四列开始的所有单元格
    for col_num in range(4, max_col + 1):
        cell = sheet.cell(row=row_num, column=col_num)
        # 将单元格的值转换为数字，如果无法转换，则跳过该单元格
        try:
            cell_value = float(cell.value)
        except (ValueError, TypeError):
            cell_value = None

        if cell_value is not None and (max_value is None or cell_value > max_value):
            max_value = cell_value
            max_col_letter = openpyxl.utils.get_column_letter(col_num)

    # 将最大值所在的单元格标记为红色
    if max_col_letter is not None:
        sheet[max_col_letter + str(row_num)].fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

# 保存修改后的Excel文件为一个新的.xlsx文件
wb.save('3.xlsx')
